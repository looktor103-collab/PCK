"""
QCMD HPV EQA Interactive Dashboard - Laboratory TH193
Creates a comprehensive, beautiful Excel dashboard for HPV EQA results 2024-2026
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# COLORS & HELPERS
# ══════════════════════════════════════════════════════════════════════════════
NAVY   = "1B3A5C"
TEAL   = "006D77"
TEAL2  = "2A9D8F"
LBLUE  = "D6EAF8"
LBLUE2 = "EBF5FB"
WHITE  = "FFFFFF"
GREEN  = "1A7A41"
DGREEN = "145A32"
LGREEN = "D5F5E3"
RED    = "922B21"
LRED   = "FADBD8"
GOLD   = "9A7D0A"
LGOLD  = "FEF9E7"
LGRAY  = "F4F6F7"
MGRAY  = "BDC3C7"
DGRAY  = "2C3E50"
ORANGE = "CA6F1E"
BLUE2  = "154360"
PURPLE = "6C3483"
LPURPLE= "E8DAEF"

def f(bold=False, sz=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)

def fl(c):
    return PatternFill("solid", fgColor=c)

def bd(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(cell, v=None, bold=False, sz=11, fc="000000", bg=None,
       ha="left", va="center", wrap=False, italic=False, nf=None, brd="thin"):
    if v is not None: cell.value = v
    cell.font = f(bold, sz, fc, italic)
    if bg: cell.fill = fl(bg)
    cell.alignment = al(ha, va, wrap)
    if brd: cell.border = bd(brd)
    if nf: cell.number_format = nf

def mc(ws, rng, v=None, bold=False, sz=11, fc="FFFFFF", bg=None,
       ha="center", va="center", wrap=False, italic=False):
    ws.merge_cells(rng)
    c0 = rng.split(":")[0]
    c = ws[c0]
    sc(c, v, bold, sz, fc, bg, ha, va, wrap, italic, brd=None)
    return c

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h

def th(ws, row, col, headers, fg=WHITE, bg=NAVY, sz=10):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col+i)
        sc(c, h, bold=True, sz=sz, fc=fg, bg=bg, ha="center", va="center", wrap=True, brd="thin")

def data_row(ws, row, col, values, bg=WHITE, sz=10, bold_first=False):
    row_bg = bg
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col+i)
        is_bold = bold_first and i == 0
        sc(c, v, bold=is_bold, sz=sz, fc=DGRAY, bg=row_bg, ha="center", va="center", brd="thin")

# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════
CHALLENGES = [
    {"year":2024,"ch":"C1","month":"May 2024","ds":191,"part":178,"ctry":30,"score":0},
    {"year":2024,"ch":"C2","month":"Oct 2024","ds":196,"part":180,"ctry":30,"score":0},
    {"year":2025,"ch":"C1","month":"May 2025","ds":256,"part":225,"ctry":29,"score":0},
    {"year":2025,"ch":"C2","month":"Oct 2025","ds":253,"part":225,"ctry":29,"score":0},
    {"year":2026,"ch":"C1","month":"May 2026","ds":225,"part":195,"ctry":31,"score":0},
]

# (year, ch, month, sno, content, genotype, status, result, dscore, ct, peer_pct)
SAMPLES = [
    (2024,"C1","May 2024","01","HPV18 (Hela)","HPV18","CORE","Positive",0,29.70,99.0),
    (2024,"C1","May 2024","02","HPV16 (Caski)","HPV16","CORE","Positive",0,26.44,99.5),
    (2024,"C1","May 2024","03","HPV18 (Hela)","HPV18","CORE","Positive",0,26.53,99.0),
    (2024,"C1","May 2024","04","HPV18 (Hela)","HPV18","CORE","Positive",0,23.51,99.0),
    (2024,"C1","May 2024","05","HPV16 (Caski)","HPV16","EDUCATIONAL","Positive",0,31.05,84.8),
    (2024,"C1","May 2024","06","Negative","Negative","CORE","Negative",0,None,99.0),
    (2024,"C2","Oct 2024","01","HPV16 (Caski)","HPV16","CORE","Positive",0,25.49,98.5),
    (2024,"C2","Oct 2024","02","HPV45 (CC10b)","HPV45","CORE","Positive",0,26.63,99.0),
    (2024,"C2","Oct 2024","03","Negative","Negative","CORE","Negative",0,None,98.0),
    (2024,"C2","Oct 2024","04","HPV45 (CC10b)","HPV45","CORE","Positive",0,26.54,98.5),
    (2024,"C2","Oct 2024","05","HPV18 (Hela)","HPV18","CORE","Positive",0,26.75,98.0),
    (2024,"C2","Oct 2024","06","HPV16+HPV18","HPV16+18","CORE","Positive",0,27.16,99.0),
    (2025,"C1","May 2025","01","HPV16 (Caski)","HPV16","CORE","Positive",0,27.26,100.0),
    (2025,"C1","May 2025","02","HPV18 (Hela)","HPV18","CORE","Positive",0,27.37,98.8),
    (2025,"C1","May 2025","03","HPV18 (Hela)","HPV18","CORE","Positive",0,23.43,99.6),
    (2025,"C1","May 2025","04","HPV16 (Caski)","HPV16","CORE","Positive",0,27.21,99.6),
    (2025,"C1","May 2025","05","Negative","Negative","CORE","Negative",0,None,98.0),
    (2025,"C1","May 2025","06","HPV18 low (Hela)","HPV18","CORE","Positive",0,30.15,94.9),
    (2025,"C2","Oct 2025","01","HPV16 (Caski)","HPV16","CORE","Positive",0,25.39,99.2),
    (2025,"C2","Oct 2025","02","HPV45 (CC10b)","HPV45","CORE","Positive",0,27.70,98.4),
    (2025,"C2","Oct 2025","03","HPV18 (Hela)","HPV18","CORE","Positive",0,26.48,98.0),
    (2025,"C2","Oct 2025","04","Negative","Negative","CORE","Negative",0,None,98.8),
    (2025,"C2","Oct 2025","05","HPV45 (CC10b)","HPV45","CORE","Positive",0,27.55,98.4),
    (2025,"C2","Oct 2025","06","HPV16 (Caski)","HPV16","CORE","Positive",0,24.91,99.2),
    (2026,"C1","May 2026","01","HPV16 (Caski)","HPV16","CORE","Positive",0,28.41,96.9),
    (2026,"C1","May 2026","02","HPV16 (Caski)","HPV16","CORE","Positive",0,25.63,99.1),
    (2026,"C1","May 2026","03","HPV16 (Caski)","HPV16","CORE","Positive",0,26.16,99.1),
    (2026,"C1","May 2026","04","HPV45 (CC10b)","HPV45","CORE","Positive",0,26.64,98.7),
    (2026,"C1","May 2026","05","HPV16 (Caski)","HPV16","CORE","Positive",0,29.99,97.3),
    (2026,"C1","May 2026","06","Negative","Negative","CORE","Negative",0,None,98.2),
]

def get_stats(year=None, ch=None):
    rows = [s for s in SAMPLES
            if (year is None or s[0]==year) and (ch is None or s[1]==ch)]
    tp = sum(1 for s in rows if s[7]=="Positive" and s[5]!="Negative")
    tn = sum(1 for s in rows if s[7]=="Negative" and s[5]=="Negative")
    fp = sum(1 for s in rows if s[7]=="Positive" and s[5]=="Negative")
    fn = sum(1 for s in rows if s[7]=="Negative" and s[5]!="Negative")
    sens = tp/(tp+fn)*100 if (tp+fn)>0 else 100.0
    spec = tn/(tn+fp)*100 if (tn+fp)>0 else 100.0
    core_peer = [s[10] for s in rows if s[6]=="CORE"]
    avg_peer = sum(core_peer)/len(core_peer) if core_peer else 0.0
    return {"tp":tp,"tn":tn,"fp":fp,"fn":fn,"sens":sens,"spec":spec,"avg_peer":avg_peer}

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_db = wb.active
ws_db.title = "Dashboard"
ws_db.sheet_view.showGridLines = False
ws_db.sheet_view.zoomScale = 90

# Column widths
for c, w in [(1,2),(2,18),(3,13),(4,14),(5,13),(6,13),(7,13),(8,14),(9,14),(10,16),(11,2)]:
    cw(ws_db, c, w)

# ── TITLE BLOCK ──
rh(ws_db, 1, 6)
rh(ws_db, 2, 54)
mc(ws_db,"B2:J2","🔬   QCMD HPV EQA Performance Dashboard",
   bold=True, sz=22, fc=WHITE, bg=NAVY, ha="center")

rh(ws_db, 3, 28)
mc(ws_db,"B3:J3",
   "QCMD HPV PreservCyt (QAV094130)  │  Laboratory TH193 - Thailand  │  Roche Cobas HPV (v4)  │  Cobas 5800/6800/8800  │  Reporting Period: 2024–2026",
   bold=False, sz=10, fc=WHITE, bg=TEAL, ha="center", italic=True)

rh(ws_db, 4, 8)

# ── KPI SECTION HEADER ──
rh(ws_db, 5, 16)
mc(ws_db,"B5:J5","  KEY PERFORMANCE INDICATORS",
   bold=True, sz=9, fc=MGRAY, bg=LGRAY, ha="left")
ws_db["B5"].font = Font(name="Arial", bold=True, size=9, color=MGRAY)

# ── KPI ROW 1 (big numbers) ──
rh(ws_db, 6, 52)
rh(ws_db, 7, 28)

# Box 1 - Total Challenges
mc(ws_db,"B6:C6","5", bold=True, sz=36, fc=WHITE, bg=NAVY, ha="center")
mc(ws_db,"B7:C7","Challenges\nCompleted", bold=False, sz=10, fc=WHITE, bg=NAVY, ha="center", wrap=True)

# Box 2 - All PASS
mc(ws_db,"D6:E6","5/5", bold=True, sz=36, fc=WHITE, bg=GREEN, ha="center")
mc(ws_db,"D7:E7","PASS\nAll Challenges", bold=True, sz=10, fc=WHITE, bg=GREEN, ha="center", wrap=True)

# Box 3 - Sensitivity
mc(ws_db,"F6:G6","100%", bold=True, sz=36, fc=WHITE, bg=TEAL, ha="center")
mc(ws_db,"F7:G7","Sensitivity\n(All Samples)", bold=False, sz=10, fc=WHITE, bg=TEAL, ha="center", wrap=True)

# Box 4 - Specificity
mc(ws_db,"H6:J6","100%", bold=True, sz=36, fc=WHITE, bg=BLUE2, ha="center")
mc(ws_db,"H7:J7","Specificity\n(All Samples)", bold=False, sz=10, fc=WHITE, bg=BLUE2, ha="center", wrap=True)

rh(ws_db, 8, 8)

# ── KPI ROW 2 (secondary stats) ──
rh(ws_db, 9, 48)
rh(ws_db,10, 8)

def kpi2(ws, rng, val, label):
    mc(ws, rng, val+"\n"+label, bold=True, sz=11, fc=NAVY, bg=LBLUE2, ha="center", wrap=True)

kpi2(ws_db,"B9:C9","30","Total Samples Tested")
kpi2(ws_db,"D9:E9","29","CORE Samples")
kpi2(ws_db,"F9:G9","1","Educational Sample")
kpi2(ws_db,"H9:J9","98.4%","Avg Peer % Correct (CORE)")

# ── ANNUAL SUMMARY SECTION HEADER ──
rh(ws_db,11,16)
mc(ws_db,"B11:J11","  ANNUAL PERFORMANCE SUMMARY — Challenge-by-Challenge",
   bold=True, sz=9, fc=MGRAY, bg=LGRAY, ha="left")

# ── Table header ──
rh(ws_db,12,26)
hdrs = ["Year","Challenge","Date","Datasets","Core Score","EQA Result","Sensitivity","Specificity","Avg Peer %"]
th(ws_db, 12, 2, hdrs, fg=WHITE, bg=NAVY, sz=10)

# ── Data rows ──
for idx, chd in enumerate(CHALLENGES):
    row = 13 + idx
    rh(ws_db, row, 20)
    stats = get_stats(chd["year"], chd["ch"])
    row_bg = WHITE if idx % 2 == 0 else LGRAY

    cells_vals = [
        (chd["year"], True, NAVY, row_bg),
        (chd["ch"], True, TEAL, row_bg),
        (chd["month"], False, DGRAY, row_bg),
        (chd["ds"], False, DGRAY, row_bg),
        (f"Score: {chd['score']}  ✓", False, GREEN, row_bg),
        ("✅  PASS", True, GREEN, LGREEN),
        (f"{stats['sens']:.1f}%", True, DGREEN, LGREEN),
        (f"{stats['spec']:.1f}%", True, DGREEN, LGREEN),
        (f"{stats['avg_peer']:.1f}%", False, BLUE2, row_bg),
    ]
    for i, (v, bold, fc, bg) in enumerate(cells_vals):
        c = ws_db.cell(row=row, column=2+i)
        sc(c, v, bold=bold, sz=10, fc=fc, bg=bg, ha="center", va="center", brd="thin")

# ── NAVIGATION ──
rh(ws_db, 18, 10)
rh(ws_db, 19, 16)
mc(ws_db,"B19:J19","  QUICK NAVIGATION — Click to jump to detailed reports",
   bold=True, sz=9, fc=MGRAY, bg=LGRAY, ha="left")

rh(ws_db, 20, 26)
nav = [
    ("B","C","2024 Detailed Results","'2024'!A1",TEAL),
    ("D","E","2025 Detailed Results","'2025'!A1",TEAL2),
    ("F","G","2026 Detailed Results","'2026'!A1",BLUE2),
    ("H","J","Genotype Analysis","'Genotype Analysis'!A1",NAVY),
]
for c1, c2, label, link, color in nav:
    rng = f"{c1}20:{c2}20"
    ws_db.merge_cells(rng)
    c = ws_db[f"{c1}20"]
    c.value = f"→  {label}"
    c.hyperlink = f"#{link}"
    c.font = Font(name="Arial", bold=True, size=11, color=WHITE, underline="single")
    c.fill = fl(color)
    c.alignment = al("center")

rh(ws_db, 21, 8)

# ── CHART DATA (hidden in rows 60+) ──
# Challenge labels and avg peer % for bar chart
chart_start = 60
labels_col = 1
peer_col = 2
th193_col = 3

ws_db.cell(60,1,"Challenge")
ws_db.cell(60,2,"Avg Peer % (CORE)")
ws_db.cell(60,3,"TH193 (100%)")

ch_labels = []
for i, chd in enumerate(CHALLENGES):
    stats = get_stats(chd["year"], chd["ch"])
    r = 61 + i
    ws_db.cell(r,1,f"{chd['year']} {chd['ch']}")
    ws_db.cell(r,2,round(stats["avg_peer"],1))
    ws_db.cell(r,3,100.0)
    ch_labels.append(f"{chd['year']} {chd['ch']}")

# Bar chart: Peer % vs TH193 by challenge
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Peer Group % Correct vs TH193 by Challenge"
chart1.y_axis.title = "% Correct"
chart1.x_axis.title = "Challenge"
chart1.y_axis.scaling.min = 80
chart1.y_axis.scaling.max = 101
chart1.height = 13
chart1.width = 22
chart1.grouping = "clustered"

cats = Reference(ws_db, min_col=1, min_row=61, max_row=65)
peer_ref = Reference(ws_db, min_col=2, min_row=60, max_row=65)
th193_ref = Reference(ws_db, min_col=3, min_row=60, max_row=65)
chart1.add_data(peer_ref, titles_from_data=True)
chart1.add_data(th193_ref, titles_from_data=True)
chart1.set_categories(cats)
chart1.series[0].graphicalProperties.solidFill = "2196F3"
chart1.series[1].graphicalProperties.solidFill = "4CAF50"
ws_db.add_chart(chart1, "B22")

# ══════════════════════════════════════════════════════════════════════════════
# HELPER: build a year detail sheet
# ══════════════════════════════════════════════════════════════════════════════
def build_year_sheet(wb, year, challenges_in_year):
    ws = wb.create_sheet(str(year))
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 95

    # Column widths: A=2, B=14, C=12, D=22, E=12, F=12, G=12, H=13, I=13, J=14, K=2
    for c, w in [(1,2),(2,14),(3,12),(4,24),(5,12),(6,12),(7,12),(8,13),(9,13),(10,16),(11,2)]:
        cw(ws, c, w)

    rh(ws, 1, 6)
    rh(ws, 2, 46)
    mc(ws,"B2:J2",f"🗓  {year} EQA Results — Laboratory TH193",
       bold=True, sz=20, fc=WHITE, bg=NAVY, ha="center")

    rh(ws, 3, 26)
    mc(ws,"B3:J3","QCMD HPV PreservCyt │ QAV094130 │ Roche Cobas HPV (v4) │ Cobas 5800/6800/8800",
       bold=False, sz=10, fc=WHITE, bg=TEAL, ha="center", italic=True)
    rh(ws, 4, 8)

    # Hyperlink back to Dashboard
    rh(ws, 5, 20)
    ws.merge_cells("B5:C5")
    c = ws["B5"]
    c.value = "← Dashboard"
    c.hyperlink = "#Dashboard!A1"
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE, underline="single")
    c.fill = fl(TEAL)
    c.alignment = al("center")

    cur_row = 7

    for ch_idx, chd in enumerate(challenges_in_year):
        ch = chd["ch"]
        stats = get_stats(year, ch)
        smpls = [s for s in SAMPLES if s[0]==year and s[1]==ch]

        # Challenge header
        ch_color = TEAL if ch=="C1" else BLUE2
        rh(ws, cur_row, 10)
        mc(ws,f"B{cur_row}:J{cur_row}","",bg=LGRAY)
        cur_row += 1

        rh(ws, cur_row, 28)
        mc(ws,f"B{cur_row}:J{cur_row}",
           f"  Challenge {ch}  ─  {chd['month']}",
           bold=True, sz=14, fc=WHITE, bg=ch_color, ha="left")
        cur_row += 1

        # Challenge meta info
        rh(ws, cur_row, 22)
        meta_items = [
            (f"Datasets: {chd['ds']}", "B"),
            (f"Participants: {chd['part']}", "D"),
            (f"Countries: {chd['ctry']}", "F"),
            (f"Core Score: {chd['score']} (Highly Satisfactory)", "H"),
        ]
        for text, col_letter in meta_items:
            c2 = f"{col_letter}{cur_row}"
            c3 = f"{chr(ord(col_letter)+1)}{cur_row}"
            ws.merge_cells(f"{c2}:{c3}")
            cell = ws[c2]
            sc(cell, text, bold=False, sz=10, fc=DGRAY, bg=LBLUE2, ha="center", va="center", brd=None)
            cell.border = bd("thin")
        cur_row += 1

        # Sensitivity/Specificity bar
        rh(ws, cur_row, 22)
        tp, tn = stats["tp"], stats["tn"]
        n_pos = sum(1 for s in smpls if s[5]!="Negative")
        n_neg = sum(1 for s in smpls if s[5]=="Negative")

        ss_items = [
            (f"Sensitivity: {stats['sens']:.1f}%  (TP={tp}/{n_pos})", LGREEN, GREEN),
            (f"Specificity: {stats['spec']:.1f}%  (TN={tn}/{n_neg})", LGREEN, GREEN),
            (f"EQA Result: ✅ PASS", LGREEN, DGREEN),
            (f"Avg Peer %: {stats['avg_peer']:.1f}%", LBLUE, BLUE2),
        ]
        col_pairs = [("B","C"),("D","E"),("F","G"),("H","J")]
        for (text, bg, fc), (c1, c2) in zip(ss_items, col_pairs):
            ws.merge_cells(f"{c1}{cur_row}:{c2}{cur_row}")
            cell = ws[f"{c1}{cur_row}"]
            sc(cell, text, bold=True, sz=10, fc=fc, bg=bg, ha="center", va="center", brd=None)
            cell.border = bd("thin")
        cur_row += 1

        rh(ws, cur_row, 8)
        cur_row += 1

        # Sample results table header
        rh(ws, cur_row, 24)
        sample_hdrs = ["Sample","Content","HPV Type","Status","Expected","TH193 Result","Det. Score","CT Value","Peer % Correct"]
        th(ws, cur_row, 2, sample_hdrs, fg=WHITE, bg=ch_color, sz=10)
        cur_row += 1

        for s_idx, s in enumerate(smpls):
            rh(ws, cur_row, 20)
            row_bg = WHITE if s_idx % 2 == 0 else LGRAY

            # Determine expected result
            expected = "Positive" if s[5]!="Negative" else "Negative"
            is_correct = s[7] == expected

            # Result coloring
            result_bg = LGREEN if is_correct else LRED
            result_fc = GREEN if is_correct else RED

            # Status coloring
            status_bg = LBLUE if s[6]=="CORE" else LGOLD
            status_fc = BLUE2 if s[6]=="CORE" else GOLD

            # Peer % coloring - low if <95%
            peer = s[10]
            peer_bg = LGOLD if peer < 95 else row_bg

            ct_val = f"{s[9]:.2f}" if s[9] else "—"
            peer_str = f"{peer:.1f}%"

            row_vals = [
                (f"HPVPRES{str(year)[2:]}{ch}-{s[3]}", False, DGRAY, row_bg),
                (s[4], False, DGRAY, row_bg),
                (s[5], True, NAVY, row_bg),
                (s[6], True, status_fc, status_bg),
                (expected, False, DGRAY, row_bg),
                (s[7], True, result_fc, result_bg),
                (f"Score: {s[8]}  ✓", False, GREEN, LGREEN),
                (ct_val, False, DGRAY, row_bg),
                (peer_str, False, BLUE2, peer_bg),
            ]
            for i, (v, bold, fc, bg) in enumerate(row_vals):
                c = ws.cell(row=cur_row, column=2+i)
                sc(c, v, bold=bold, sz=10, fc=fc, bg=bg, ha="center", va="center", brd="thin")
            cur_row += 1

        # Add CT chart for this challenge (hidden data area)
        chart_base_row = 60 + ch_idx * 15
        ws.cell(chart_base_row, 1, "Sample")
        ws.cell(chart_base_row, 2, "CT Value")
        ws.cell(chart_base_row, 3, "Peer %")

        for i, s in enumerate(smpls):
            ws.cell(chart_base_row+1+i, 1, f"S{s[3]}")
            ws.cell(chart_base_row+1+i, 2, s[9] if s[9] else None)
            ws.cell(chart_base_row+1+i, 3, s[10])

        n_smpls = len(smpls)
        chart_ct = BarChart()
        chart_ct.type = "col"
        chart_ct.style = 10
        chart_ct.title = f"{year} {ch} — Cycle Threshold (CT) Values"
        chart_ct.y_axis.title = "CT Value"
        chart_ct.x_axis.title = "Sample"
        chart_ct.height = 10
        chart_ct.width = 18
        chart_ct.y_axis.scaling.min = 20
        chart_ct.y_axis.scaling.max = 35

        cats_ct = Reference(ws, min_col=1, min_row=chart_base_row+1, max_row=chart_base_row+n_smpls)
        ct_data = Reference(ws, min_col=2, min_row=chart_base_row, max_row=chart_base_row+n_smpls)
        chart_ct.add_data(ct_data, titles_from_data=True)
        chart_ct.set_categories(cats_ct)
        chart_ct.series[0].graphicalProperties.solidFill = "1A5276"

        # Peer % line chart overlay (LineChart cannot be combined easily, skip for now)
        ws.add_chart(chart_ct, f"B{cur_row+1}")
        cur_row += 17  # space for chart

    # Freeze top rows
    ws.freeze_panes = "B7"
    return ws

# Build year sheets
year_challenges = {
    2024: [c for c in CHALLENGES if c["year"]==2024],
    2025: [c for c in CHALLENGES if c["year"]==2025],
    2026: [c for c in CHALLENGES if c["year"]==2026],
}
for y in [2024, 2025, 2026]:
    build_year_sheet(wb, y, year_challenges[y])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: GENOTYPE ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
ws_geno = wb.create_sheet("Genotype Analysis")
ws_geno.sheet_view.showGridLines = False

for c, w in [(1,2),(2,20),(3,12),(4,12),(5,12),(6,12),(7,14),(8,14),(9,14),(10,2)]:
    cw(ws_geno, c, w)

rh(ws_geno, 1, 6)
rh(ws_geno, 2, 46)
mc(ws_geno,"B2:I2","🧬  HPV Genotype Analysis — Laboratory TH193",
   bold=True, sz=20, fc=WHITE, bg=NAVY, ha="center")
rh(ws_geno, 3, 24)
mc(ws_geno,"B3:I3","Breakdown of HPV genotypes across all EQA challenges  |  2024–2026",
   bold=False, sz=10, fc=WHITE, bg=TEAL, ha="center", italic=True)
rh(ws_geno, 4, 8)

# Navigation back
rh(ws_geno, 5, 20)
ws_geno.merge_cells("B5:C5")
c = ws_geno["B5"]
c.value = "← Dashboard"
c.hyperlink = "#Dashboard!A1"
c.font = Font(name="Arial", bold=True, size=10, color=WHITE, underline="single")
c.fill = fl(TEAL)
c.alignment = al("center")

rh(ws_geno, 6, 8)

# ── GENOTYPE FREQUENCY TABLE ──
rh(ws_geno, 7, 16)
mc(ws_geno,"B7:I7","  GENOTYPE FREQUENCY — All Challenges (2024–2026)",
   bold=True, sz=9, fc=MGRAY, bg=LGRAY, ha="left")

rh(ws_geno, 8, 24)
geno_hdrs = ["HPV Type","Samples (All)","CORE","EDUCATIONAL","Avg CT","Min CT","Max CT","Avg Peer %"]
th(ws_geno, 8, 2, geno_hdrs, fg=WHITE, bg=NAVY, sz=10)

genotypes = ["HPV16","HPV18","HPV45","HPV16+18","Negative"]
geno_colors = {
    "HPV16":   (LBLUE,   BLUE2),
    "HPV18":   (LGREEN,  GREEN),
    "HPV45":   (LPURPLE, PURPLE),
    "HPV16+18":(LGOLD,   GOLD),
    "Negative":(LGRAY,   DGRAY),
}

for idx, gtype in enumerate(genotypes):
    row = 9 + idx
    rh(ws_geno, row, 20)
    smpls_g = [s for s in SAMPLES if s[5]==gtype]
    core_g  = [s for s in smpls_g if s[6]=="CORE"]
    edu_g   = [s for s in smpls_g if s[6]=="EDUCATIONAL"]
    ct_vals = [s[9] for s in smpls_g if s[9] is not None]
    peer_core = [s[10] for s in core_g]

    avg_ct = f"{sum(ct_vals)/len(ct_vals):.2f}" if ct_vals else "—"
    min_ct = f"{min(ct_vals):.2f}" if ct_vals else "—"
    max_ct = f"{max(ct_vals):.2f}" if ct_vals else "—"
    avg_peer = f"{sum(peer_core)/len(peer_core):.1f}%" if peer_core else "—"

    bg, fc = geno_colors.get(gtype, (WHITE, DGRAY))
    row_bg = bg if idx % 2 == 0 else WHITE

    vals = [gtype, len(smpls_g), len(core_g), len(edu_g), avg_ct, min_ct, max_ct, avg_peer]
    for i, v in enumerate(vals):
        c2 = ws_geno.cell(row=row, column=2+i)
        is_label = i == 0
        sc(c2, v, bold=is_label, sz=10, fc=fc if is_label else DGRAY,
           bg=bg if is_label else (LGRAY if idx%2==1 else WHITE),
           ha="center", va="center", brd="thin")

# ── GENOTYPE × CHALLENGE MATRIX ──
rh(ws_geno, 15, 8)
rh(ws_geno, 16, 16)
mc(ws_geno,"B16:I16","  GENOTYPE × CHALLENGE MATRIX — Sample count per challenge",
   bold=True, sz=9, fc=MGRAY, bg=LGRAY, ha="left")

rh(ws_geno, 17, 24)
matrix_hdrs = ["HPV Type","2024 C1","2024 C2","2025 C1","2025 C2","2026 C1","Total","Sensitivity"]
th(ws_geno, 17, 2, matrix_hdrs, fg=WHITE, bg=NAVY, sz=10)

for idx, gtype in enumerate(genotypes):
    row = 18 + idx
    rh(ws_geno, row, 20)
    bg, fc = geno_colors.get(gtype, (WHITE, DGRAY))

    counts = []
    for chd in CHALLENGES:
        cnt = sum(1 for s in SAMPLES if s[0]==chd["year"] and s[1]==chd["ch"] and s[5]==gtype)
        counts.append(cnt)

    total = sum(counts)
    # Sensitivity = all positive genotypes were detected (only applies to non-Negative)
    sens_str = "100%" if gtype != "Negative" and total > 0 else ("N/A" if gtype=="Negative" else "—")

    vals = [gtype] + counts + [total, sens_str]
    for i, v in enumerate(vals):
        c2 = ws_geno.cell(row=row, column=2+i)
        is_label = i == 0
        sc(c2, v, bold=is_label, sz=10, fc=fc if is_label else DGRAY,
           bg=bg if is_label else (LGRAY if idx%2==1 else WHITE),
           ha="center", va="center", brd="thin")

# ── GENOTYPE CHART DATA (hidden) ──
for i, (gtype, chd) in enumerate(
    [(g, chd) for chd in CHALLENGES for g in ["HPV16","HPV18","HPV45","HPV16+18","Negative"]]
):
    pass  # skip for now, chart data below

chart_row = 60
ws_geno.cell(chart_row, 1, "Genotype")
ws_geno.cell(chart_row, 2, "# Samples")
for i, gtype in enumerate(genotypes):
    n = sum(1 for s in SAMPLES if s[5]==gtype)
    ws_geno.cell(chart_row+1+i, 1, gtype)
    ws_geno.cell(chart_row+1+i, 2, n)

chart_geno = BarChart()
chart_geno.type = "col"
chart_geno.style = 10
chart_geno.title = "HPV Genotype Distribution — All Challenges"
chart_geno.y_axis.title = "Number of Samples"
chart_geno.x_axis.title = "HPV Genotype"
chart_geno.height = 13
chart_geno.width = 20

cats_g = Reference(ws_geno, min_col=1, min_row=chart_row+1, max_row=chart_row+len(genotypes))
data_g = Reference(ws_geno, min_col=2, min_row=chart_row, max_row=chart_row+len(genotypes))
chart_geno.add_data(data_g, titles_from_data=True)
chart_geno.set_categories(cats_g)
chart_geno.series[0].graphicalProperties.solidFill = "006D77"
ws_geno.add_chart(chart_geno, "B25")

# ── SENSITIVITY & SPECIFICITY SECTION ──
# Put this after the chart (row 41+)
rh(ws_geno, 40, 16)
mc(ws_geno,"B40:I40","  SENSITIVITY & SPECIFICITY ANALYSIS — By Challenge",
   bold=True, sz=9, fc=MGRAY, bg=LGRAY, ha="left")

rh(ws_geno, 41, 24)
ss_hdrs = ["Challenge","Month","True Positive","False Negative","Sensitivity","True Negative","False Positive","Specificity"]
th(ws_geno, 41, 2, ss_hdrs, fg=WHITE, bg=NAVY, sz=10)

for idx, chd in enumerate(CHALLENGES):
    row = 42 + idx
    rh(ws_geno, row, 20)
    stats = get_stats(chd["year"], chd["ch"])
    row_bg = WHITE if idx%2==0 else LGRAY

    vals = [
        f"{chd['year']} {chd['ch']}",
        chd["month"],
        stats["tp"], stats["fn"],
        f"{stats['sens']:.1f}%",
        stats["tn"], stats["fp"],
        f"{stats['spec']:.1f}%",
    ]
    for i, v in enumerate(vals):
        c2 = ws_geno.cell(row=row, column=2+i)
        if i in [4, 7]:  # sensitivity/specificity columns
            sc(c2, v, bold=True, sz=10, fc=GREEN, bg=LGREEN, ha="center", va="center", brd="thin")
        else:
            sc(c2, v, bold=(i==0), sz=10, fc=DGRAY if i>0 else NAVY,
               bg=row_bg, ha="center", va="center", brd="thin")

ws_geno.freeze_panes = "B8"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: RAW DATA
# ══════════════════════════════════════════════════════════════════════════════
ws_raw = wb.create_sheet("Raw Data")
ws_raw.sheet_view.showGridLines = False

for c, w in [(1,6),(2,10),(3,12),(4,26),(5,12),(6,16),(7,12),(8,14),(9,12),(10,10),(11,12),(12,2)]:
    cw(ws_raw, c, w)

rh(ws_raw, 1, 6)
rh(ws_raw, 2, 40)
mc(ws_raw,"A2:K2","📋  Raw Sample Data — All Challenges 2024–2026  |  Laboratory TH193",
   bold=True, sz=16, fc=WHITE, bg=NAVY, ha="center")
rh(ws_raw, 3, 8)

# Table header
rh(ws_raw, 4, 24)
raw_hdrs = ["Year","Challenge","Month","Sample Code","Genotype","Content","Status","TH193 Result","Det. Score","CT Value","Peer %"]
th(ws_raw, 4, 1, raw_hdrs, fg=WHITE, bg=NAVY, sz=10)

for idx, s in enumerate(SAMPLES):
    row = 5 + idx
    rh(ws_raw, row, 18)
    row_bg = WHITE if idx%2==0 else LGRAY

    sample_code = f"HPVPRES{str(s[0])[2:]}{s[1]}-{s[3]}"
    ct_str = f"{s[9]:.2f}" if s[9] else "—"

    vals = [s[0], s[1], s[2], sample_code, s[5], s[4], s[6], s[7], s[8], ct_str, f"{s[10]:.1f}%"]
    for i, v in enumerate(vals):
        c2 = ws_raw.cell(row=row, column=1+i)
        # Color result column
        if i == 7:  # TH193 Result
            sc(c2, v, bold=True, sz=10, fc=GREEN, bg=LGREEN, ha="center", va="center", brd="thin")
        elif i == 6:  # Status
            st_bg = LBLUE if v=="CORE" else LGOLD
            st_fc = BLUE2 if v=="CORE" else GOLD
            sc(c2, v, bold=False, sz=10, fc=st_fc, bg=st_bg, ha="center", va="center", brd="thin")
        elif i == 10:  # Peer %
            peer_val = s[10]
            p_bg = LGOLD if peer_val < 95 else row_bg
            sc(c2, v, bold=False, sz=10, fc=DGRAY, bg=p_bg, ha="center", va="center", brd="thin")
        else:
            sc(c2, v, bold=(i==0), sz=10, fc=NAVY if i==0 else DGRAY,
               bg=row_bg, ha="center", va="center", brd="thin")

ws_raw.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET ORDER & SAVE
# ══════════════════════════════════════════════════════════════════════════════
# Reorder sheets
sheet_order = ["Dashboard","2024","2025","2026","Genotype Analysis","Raw Data"]
for i, name in enumerate(sheet_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

output_path = "/sessions/intelligent-amazing-feynman/mnt/outputs/QCMD_HPV_EQA_Dashboard_TH193.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
