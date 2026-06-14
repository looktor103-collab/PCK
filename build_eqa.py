from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
import openpyxl.chart as chart_module

wb = Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
NAVY    = "1B3A6B"
TEAL    = "0D7377"
GREEN   = "14A044"
GOLD    = "F5A623"
RED     = "D0021B"
WHITE   = "FFFFFF"
LGRAY   = "F0F4F8"
DGRAY   = "4A5568"
MINT    = "E8F5E9"
SKYBLUE = "E3F2FD"
AMBER   = "FFF8E1"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def border(color="CCCCCC", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(ws, row, col, value, bg=NAVY, fg=WHITE, sz=12, bold=True, span=None, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h_align, "center", True)
    c.border = border()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def style_cell(ws, row, col, value, bg=WHITE, fg="000000", sz=11,
               bold=False, h_align="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h_align, "center", wrap)
    c.border = border()
    return c

# ════════════════════════════════════════════════════════════════════════════
# DATA
# ════════════════════════════════════════════════════════════════════════════
summary_data = [
    # (year_th, year_en, lab, round, report_no, date, score, max_score, pct, assessment, method)
    (2566, 2023, "00115", 1, "66-002-113",  "17 เม.ย. 2566", 10, 10, 100, "Excellent",
     "LightCycler Multiplex RNA virus master (TIB Molbiol) + DirectDetect (COYOTE)"),
    (2566, 2023, "00115", 2, "66-004-108",  "14 ก.ย. 2566",  10, 10, 100, "Excellent",
     "LightCycler® Multiplex RNA Virus Master (TIB MOLBIOL) + Cobas® SARS-CoV-2 Test (ROCHE)"),
    (2567, 2024, "00096", 1, "SARS 67-02-113","11 เม.ย. 2567", 10, 10, 100, "Excellent",
     "SANSURE BIOTECH Nucleic Acid Diagnostic Kit – iPonatic III"),
    (2567, 2024, "00096", 2, "SARS 67-04-106","9 ก.ย. 2567",  10, 10, 100, "Excellent",
     "SANSURE BIOTECH Nucleic Acid Diagnostic Kit – iPonatic III"),
    (2568, 2025, "00053", 1, "SARS 68-02-067","10 เม.ย. 2568", 5, 5, 100, "Excellent",
     "SANSURE BIOTECH Nucleic Acid Diagnostic Kit – iPotanic"),
    (2568, 2025, "00053", 2, "SARS 68-04-066","8 ก.ย. 2568",  5, 5, 100, "Excellent",
     "SANSURE BIOTECH Nucleic Acid Diagnostic Kit – iPonatic"),
    (2569, 2026, "00072", 1, "SARS 69-02-070","15 พ.ค. 2569", 5, 5, 100, "Excellent",
     "SANSURE BIOTECH Nucleic Acid Diagnostic Kit + Sample release Reagent"),
]

sample_data = {
    (2566,1): [("nCoV-1G","Detected","Detected","✓"),
               ("nCoV-2G","Detected","Detected","✓"),
               ("nCoV-3G","Not detected","Not detected","✓"),
               ("nCoV-4G","Detected","Detected","✓"),
               ("nCoV-5G","Detected","Detected","✓")],
    (2566,2): [("nCoV-1H","Detected","Detected","✓"),
               ("nCoV-2H","Detected","Detected","✓"),
               ("nCoV-3H","Detected","Detected","✓"),
               ("nCoV-4H","Detected","Detected","✓"),
               ("nCoV-5H","Not detected","Not detected","✓")],
    (2567,1): [("SARS-1/2024","Detected","Detected","✓"),
               ("SARS-2/2024","Detected","Detected","✓"),
               ("SARS-3/2024","Detected","Detected","✓"),
               ("SARS-4/2024","Detected","Detected","✓"),
               ("SARS-5/2024","Detected","Detected","✓")],
    (2567,2): [("SARS-6/2024","Detected","Detected","✓"),
               ("SARS-7/2024","Detected","Detected","✓"),
               ("SARS-8/2024","Not detected","Not detected","✓"),
               ("SARS-9/2024","Detected","Detected","✓"),
               ("SARS-10/2024","Not detected","Not detected","✓")],
    (2568,1): [("SARS-1/2025","Detected","Detected","✓"),
               ("SARS-2/2025","Not detected","Not detected","✓"),
               ("SARS-3/2025","Detected","Detected","✓"),
               ("SARS-4/2025","Detected","Detected","✓"),
               ("SARS-5/2025","Detected","Detected","✓")],
    (2568,2): [("SARS-6/2025","Detected","Detected","✓"),
               ("SARS-7/2025","Not detected","Not detected","✓"),
               ("SARS-8/2025","Detected","Detected","✓"),
               ("SARS-9/2025","Not detected","Not detected","✓"),
               ("SARS-10/2025","Detected","Detected","✓")],
    (2569,1): [("SARS-1/2026","Negative","Negative","✓"),
               ("SARS-2/2026","Negative","Negative","✓"),
               ("SARS-3/2026","SARS-CoV-2","SARS-CoV-2","✓"),
               ("SARS-4/2026","SARS-CoV-2","SARS-CoV-2","✓"),
               ("SARS-5/2026","SARS-CoV-2","SARS-CoV-2","✓")],
}

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "📊 Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.row_dimensions[1].height = 8

# ── Title banner ─────────────────────────────────────────────────────────
for r in range(2, 8):
    ws_dash.row_dimensions[r].height = 18
ws_dash.merge_cells("B2:L7")
c = ws_dash["B2"]
c.value = "EQA DASHBOARD\nแผนทดสอบความชำนาญการตรวจสารพันธุกรรมไวรัส SARS-CoV-2\nด้วยวิธีทางอณูชีววิทยา"
c.font = Font(name="Arial", bold=True, size=20, color=WHITE)
c.fill = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Sub-title strip
ws_dash.row_dimensions[8].height = 22
ws_dash.merge_cells("B8:L8")
c8 = ws_dash["B8"]
c8.value = "สถาบันวิจัยวิทยาศาสตร์สาธารณสุข  กรมวิทยาศาสตร์การแพทย์  |  ปี พ.ศ. 2566 – 2569"
c8.font = Font(name="Arial", bold=False, size=11, color=WHITE)
c8.fill = fill(TEAL)
c8.alignment = align("center","center")

# ── KPI Cards row ────────────────────────────────────────────────────────
ws_dash.row_dimensions[10].height = 16
ws_dash.row_dimensions[11].height = 40
ws_dash.row_dimensions[12].height = 28
ws_dash.row_dimensions[13].height = 22

kpis = [
    ("B", "จำนวนปีที่เข้าร่วม", "4 ปี",  TEAL,  "2566–2569"),
    ("E", "ครั้งทดสอบทั้งหมด",  "7 ครั้ง", NAVY, "รวม 7 rounds"),
    ("H", "คะแนนเฉลี่ย",        "100%",   GREEN,  "ผ่านทุกครั้ง"),
    ("K", "ผลประเมินโดยรวม",    "Excellent","F5A623","ระดับดีเยี่ยม"),
]

for col_letter, label, value, bg, sub in kpis:
    col = ord(col_letter) - ord('A') + 1
    ws_dash.merge_cells(start_row=11, start_column=col,
                        end_row=11, end_column=col+1)
    ws_dash.merge_cells(start_row=12, start_column=col,
                        end_row=12, end_column=col+1)
    ws_dash.merge_cells(start_row=13, start_column=col,
                        end_row=13, end_column=col+1)

    c = ws_dash.cell(row=11, column=col, value=label)
    c.font = Font(name="Arial", size=10, color=WHITE, bold=True)
    c.fill = fill(bg)
    c.alignment = align("center","center")

    c2 = ws_dash.cell(row=12, column=col, value=value)
    c2.font = Font(name="Arial", size=22, color=WHITE, bold=True)
    c2.fill = fill(bg)
    c2.alignment = align("center","center")

    c3 = ws_dash.cell(row=13, column=col, value=sub)
    c3.font = Font(name="Arial", size=9, color=WHITE)
    c3.fill = fill(bg)
    c3.alignment = align("center","center")

# ── Summary Table ────────────────────────────────────────────────────────
ws_dash.row_dimensions[15].height = 20
ws_dash.merge_cells("B15:L15")
title15 = ws_dash["B15"]
title15.value = "📋  สรุปผลการทดสอบความชำนาญ EQA ทุกปี"
title15.font = Font(name="Arial", bold=True, size=13, color=WHITE)
title15.fill = fill(NAVY)
title15.alignment = align("left","center")

headers = ["ปี (พ.ศ.)", "ปี (ค.ศ.)", "รหัสห้องปฏิบัติการ",
           "ครั้งที่", "หมายเลขรายงาน", "วันที่ออกรายงาน",
           "คะแนนที่ได้", "คะแนนเต็ม", "ร้อยละ", "ผลประเมิน"]
header_cols = list(range(2, 12))  # B-K
ws_dash.row_dimensions[16].height = 22
for i, (col, hdr) in enumerate(zip(header_cols, headers)):
    c = ws_dash.cell(row=16, column=col, value=hdr)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = fill(TEAL)
    c.alignment = align("center","center", True)
    c.border = border()

row_colors = [LGRAY, WHITE]
for ri, sd in enumerate(summary_data):
    r = 17 + ri
    ws_dash.row_dimensions[r].height = 18
    bg = row_colors[ri % 2]
    vals = [sd[0], sd[1], sd[2], f"ครั้งที่ {sd[3]}",
            sd[4], sd[5], sd[6], sd[7], f"{sd[8]}%", sd[9]]
    for ci, (col, val) in enumerate(zip(header_cols, vals)):
        c = ws_dash.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.fill = fill(bg)
        c.alignment = align("center","center")
        c.border = border()
        # Assessment coloring
        if ci == 9:
            c.font = Font(name="Arial", size=10, bold=True, color=GREEN)

# ── Legend / criteria ────────────────────────────────────────────────────
r_leg = 25
ws_dash.row_dimensions[r_leg].height = 18
ws_dash.merge_cells(f"B{r_leg}:L{r_leg}")
leg_title = ws_dash.cell(row=r_leg, column=2,
                          value="เกณฑ์การประเมินผล EQA")
leg_title.font = Font(name="Arial", bold=True, size=11, color=WHITE)
leg_title.fill = fill(DGRAY)
leg_title.alignment = align("left","center")

criteria = [
    ("🏆 Excellent", "ร้อยละ 100",     GREEN,  MINT),
    ("⚠️ Borderline","ร้อยละ 80–99",    GOLD,   AMBER),
    ("❌ Unacceptable","น้อยกว่าร้อยละ 80", RED, "FFEBEE"),
]
for ci, (label, pct, fg, bg) in enumerate(criteria):
    col = 2 + ci * 3
    ws_dash.merge_cells(start_row=r_leg+1, start_column=col,
                        end_row=r_leg+1, end_column=col+2)
    cell = ws_dash.cell(row=r_leg+1, column=col,
                         value=f"{label}  |  {pct}")
    cell.font = Font(name="Arial", bold=True, size=10, color=fg)
    cell.fill = fill(bg)
    cell.alignment = align("center","center")
    cell.border = border()

# column widths dashboard
col_widths_dash = {1:2, 2:10, 3:8, 4:16, 5:10,
                   6:18, 7:16, 8:10, 9:10, 10:8, 11:14, 12:2}
for col, w in col_widths_dash.items():
    ws_dash.column_dimensions[get_column_letter(col)].width = w

# ── Bar chart on dashboard ───────────────────────────────────────────────
# Use hidden helper data for chart
chart_row = 28
ws_dash.cell(row=chart_row,   column=2, value="ปี")
ws_dash.cell(row=chart_row,   column=3, value="ร้อยละ")
years_chart = ["2566 R1","2566 R2","2567 R1","2567 R2",
               "2568 R1","2568 R2","2569 R1"]
pcts_chart  = [100,100,100,100,100,100,100]
for i,(yr,pc) in enumerate(zip(years_chart,pcts_chart)):
    ws_dash.cell(row=chart_row+1+i, column=2, value=yr)
    ws_dash.cell(row=chart_row+1+i, column=3, value=pc)

bar = BarChart()
bar.type = "col"
bar.title = "ร้อยละความถูกต้องแต่ละครั้งทดสอบ"
bar.y_axis.title = "ร้อยละ (%)"
bar.x_axis.title = "รอบทดสอบ"
bar.style = 10
bar.width = 22
bar.height = 12

data_ref = Reference(ws_dash, min_col=3, min_row=chart_row,
                     max_row=chart_row+7)
cats = Reference(ws_dash, min_col=2, min_row=chart_row+1,
                 max_row=chart_row+7)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = "14A044"
ws_dash.add_chart(bar, "F25")

# ════════════════════════════════════════════════════════════════════════════
# YEARLY SHEETS
# ════════════════════════════════════════════════════════════════════════════
year_config = {
    2566: {"title_color": "1B3A6B", "strip_color": "0D7377", "bg": "E8F5E9"},
    2567: {"title_color": "0D7377", "strip_color": "1B3A6B", "bg": "E3F2FD"},
    2568: {"title_color": "6A1B9A", "strip_color": "AB47BC", "bg": "F3E5F5"},
    2569: {"title_color": "BF360C", "strip_color": "E64A19", "bg": "FBE9E7"},
}

for year_th, cfg in year_config.items():
    year_en = year_th - 543
    year_rows = [sd for sd in summary_data if sd[0] == year_th]
    if not year_rows:
        continue

    ws = wb.create_sheet(title=f"📅 ปี {year_th}")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8

    tc = cfg["title_color"]
    sc = cfg["strip_color"]
    row_bg = cfg["bg"]

    # Banner
    for r in range(2, 7):
        ws.row_dimensions[r].height = 18
    ws.merge_cells("B2:M6")
    banner = ws["B2"]
    banner.value = (f"EQA ปี พ.ศ. {year_th}  (ค.ศ. {year_en})\n"
                    f"แผนทดสอบความชำนาญการตรวจสารพันธุกรรมไวรัส SARS-CoV-2\n"
                    f"ด้วยวิธีทางอณูชีววิทยา")
    banner.font = Font(name="Arial", bold=True, size=18, color=WHITE)
    banner.fill = fill(tc)
    banner.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)

    ws.row_dimensions[7].height = 20
    ws.merge_cells("B7:M7")
    sub = ws["B7"]
    sub.value = (f"รหัสห้องปฏิบัติการ : {year_rows[0][2]}  |  "
                 f"ผู้รับรอง : สถาบันวิจัยวิทยาศาสตร์สาธารณสุข กรมวิทยาศาสตร์การแพทย์")
    sub.font = Font(name="Arial", size=11, color=WHITE)
    sub.fill = fill(sc)
    sub.alignment = align("center","center")

    # KPI
    ws.row_dimensions[9].height = 38
    ws.row_dimensions[10].height = 22
    for ri, sd in enumerate(year_rows):
        col = 2 + ri * 5
        ws.merge_cells(start_row=9, start_column=col,
                       end_row=9, end_column=col+3)
        kpi_cell = ws.cell(row=9, column=col,
                           value=f"ครั้งที่ {sd[3]}  |  {sd[9]}  |  {sd[8]}%")
        kpi_cell.font = Font(name="Arial", bold=True, size=13, color=WHITE)
        kpi_cell.fill = fill(tc)
        kpi_cell.alignment = align("center","center")

        ws.merge_cells(start_row=10, start_column=col,
                       end_row=10, end_column=col+3)
        sub2 = ws.cell(row=10, column=col, value=sd[5])
        sub2.font = Font(name="Arial", size=10, color=WHITE)
        sub2.fill = fill(sc)
        sub2.alignment = align("center","center")

    # Round tables
    current_row = 12
    for sd in year_rows:
        round_no = sd[3]
        samples = sample_data.get((year_th, round_no), [])

        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"B{current_row}:M{current_row}")
        rnd_title = ws.cell(row=current_row, column=2,
                             value=f"📌  ครั้งที่ {round_no}  |  หมายเลขรายงาน : {sd[4]}  |  วันที่ออกรายงาน : {sd[5]}")
        rnd_title.font = Font(name="Arial", bold=True, size=12, color=WHITE)
        rnd_title.fill = fill(tc)
        rnd_title.alignment = align("left","center")

        # sub-header
        current_row += 1
        sub_headers = ["รหัสวัตถุทดสอบ","ค่ากำหนด",
                        "ผลการทดสอบของสมาชิก","ผลการเปรียบเทียบ"]
        sub_cols = [2, 5, 8, 12]
        sub_spans = [3, 3, 4, 2]
        ws.row_dimensions[current_row].height = 20
        for hdr, col, span in zip(sub_headers, sub_cols, sub_spans):
            ws.merge_cells(start_row=current_row, start_column=col,
                           end_row=current_row, end_column=col+span-1)
            c = ws.cell(row=current_row, column=col, value=hdr)
            c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
            c.fill = fill(sc)
            c.alignment = align("center","center")
            c.border = border()

        # Sample rows
        for si, (samp_id, expected, result, match) in enumerate(samples):
            current_row += 1
            bg_s = LGRAY if si % 2 == 0 else WHITE
            ws.row_dimensions[current_row].height = 18

            ws.merge_cells(start_row=current_row, start_column=2,
                           end_row=current_row, end_column=4)
            c = ws.cell(row=current_row, column=2, value=samp_id)
            c.font = Font(name="Arial", size=10); c.fill = fill(bg_s)
            c.alignment = align("center","center"); c.border = border()

            ws.merge_cells(start_row=current_row, start_column=5,
                           end_row=current_row, end_column=7)
            c2 = ws.cell(row=current_row, column=5, value=expected)
            c2.font = Font(name="Arial", size=10); c2.fill = fill(bg_s)
            c2.alignment = align("center","center"); c2.border = border()

            ws.merge_cells(start_row=current_row, start_column=8,
                           end_row=current_row, end_column=11)
            c3 = ws.cell(row=current_row, column=8, value=result)
            c3.font = Font(name="Arial", size=10); c3.fill = fill(bg_s)
            c3.alignment = align("center","center"); c3.border = border()

            ws.merge_cells(start_row=current_row, start_column=12,
                           end_row=current_row, end_column=13)
            c4 = ws.cell(row=current_row, column=12, value="✅ ถูกต้อง")
            c4.font = Font(name="Arial", bold=True, size=10, color=GREEN)
            c4.fill = fill(MINT); c4.alignment = align("center","center")
            c4.border = border()

        # Score row
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=11)
        c_score = ws.cell(row=current_row, column=2,
                          value=f"คะแนนที่ได้ : {sd[6]} / {sd[7]}  |  ร้อยละ {sd[8]}%")
        c_score.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        c_score.fill = fill(TEAL)
        c_score.alignment = align("center","center")

        ws.merge_cells(start_row=current_row, start_column=12,
                       end_row=current_row, end_column=13)
        c_exc = ws.cell(row=current_row, column=12, value="🏆 Excellent")
        c_exc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        c_exc.fill = fill(GREEN)
        c_exc.alignment = align("center","center")

        # Method row
        current_row += 1
        ws.row_dimensions[current_row].height = 30
        ws.merge_cells(f"B{current_row}:M{current_row}")
        c_meth = ws.cell(row=current_row, column=2,
                          value=f"น้ำยา Real-time RT-PCR : {sd[10]}")
        c_meth.font = Font(name="Arial", size=9, italic=True, color=DGRAY)
        c_meth.fill = fill(row_bg)
        c_meth.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        current_row += 2

    # Column widths
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["B"].width = 15

# ════════════════════════════════════════════════════════════════════════════
# SHEET – ข้อมูลทั้งหมด (raw data for filtering)
# ════════════════════════════════════════════════════════════════════════════
ws_all = wb.create_sheet("📋 ข้อมูลทั้งหมด")
ws_all.sheet_view.showGridLines = False
ws_all.row_dimensions[1].height = 8

ws_all.merge_cells("B2:L4")
all_title = ws_all["B2"]
all_title.value = "ตารางข้อมูล EQA SARS-CoV-2 ทั้งหมด (2566–2569)"
all_title.font = Font(name="Arial", bold=True, size=16, color=WHITE)
all_title.fill = fill(NAVY)
all_title.alignment = align("center","center")

hdrs_all = ["ปี (พ.ศ.)", "ปี (ค.ศ.)", "รหัสห้องปฏิบัติการ",
            "ครั้งที่", "หมายเลขรายงาน", "วันที่",
            "คะแนน", "คะแนนเต็ม", "ร้อยละ",
            "ผลประเมิน", "น้ำยา"]
ws_all.row_dimensions[6].height = 22
for ci, hdr in enumerate(hdrs_all):
    c = ws_all.cell(row=6, column=2+ci, value=hdr)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = align("center","center", True)
    c.border = border()

for ri, sd in enumerate(summary_data):
    r = 7 + ri
    ws_all.row_dimensions[r].height = 18
    bg = LGRAY if ri % 2 == 0 else WHITE
    row_vals = [sd[0], sd[1], sd[2], sd[3], sd[4], sd[5],
                sd[6], sd[7], f"{sd[8]}%", sd[9], sd[10]]
    for ci, val in enumerate(row_vals):
        c = ws_all.cell(row=r, column=2+ci, value=val)
        c.font = Font(name="Arial", size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=(ci == 10))
        c.border = border()
        if ci == 9:
            c.font = Font(name="Arial", size=10, bold=True, color=GREEN)

# col widths all sheet
col_w_all = {2:10, 3:8, 4:16, 5:8, 6:18, 7:14,
             8:10, 9:10, 10:8, 11:14, 12:40}
for col, w in col_w_all.items():
    ws_all.column_dimensions[get_column_letter(col)].width = w

# ── Re-order sheets ──────────────────────────────────────────────────────
wb._sheets = [ws_dash] + [s for s in wb._sheets
                            if s.title.startswith("📅")] + [ws_all]

# ── Save ─────────────────────────────────────────────────────────────────
out_path = "/sessions/gifted-magical-meitner/mnt/outputs/EQA_SARS-CoV-2_Dashboard.xlsx"
wb.save(out_path)
print("DONE:", out_path)
